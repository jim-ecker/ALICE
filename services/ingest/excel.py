from __future__ import annotations

import hashlib
from pathlib import Path

from services.ingest.models import Chunk, ChunkProvenance, SourceDocument

_ROWS_PER_CHUNK = 30


def _find_header_row(rows: list[tuple]) -> int:
    """Return index of first row with at least 25% non-None cells."""
    for i, row in enumerate(rows):
        non_empty = sum(1 for c in row if c is not None)
        if non_empty >= max(1, len(row) // 4):
            return i
    return 0


def chunk_excel(path: Path, source_url: str | None = None) -> tuple[SourceDocument, list[Chunk]]:
    """Parse an Excel workbook into Chunk objects, one chunk per row group per sheet."""
    import openpyxl

    raw = path.read_bytes()
    doc_id = hashlib.sha256(raw).hexdigest()
    source = SourceDocument(
        id=doc_id,
        path=path,
        doc_type=path.suffix.lstrip(".").lower(),
        source_url=source_url,
        metadata={},
    )

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    chunks: list[Chunk] = []

    for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        if not rows:
            continue

        header_idx = _find_header_row(rows)
        header = rows[header_idx]

        # Only keep columns that have a non-None header
        col_indices = [i for i, h in enumerate(header) if h is not None]
        if not col_indices:
            continue

        header_cells = [str(header[i]) for i in col_indices]

        # Data rows after the header, skipping entirely-empty rows
        data_rows = [
            row for row in rows[header_idx + 1:]
            if any(row[i] is not None for i in col_indices)
        ]
        if not data_rows:
            continue

        for group_start in range(0, len(data_rows), _ROWS_PER_CHUNK):
            group = data_rows[group_start: group_start + _ROWS_PER_CHUNK]
            lines = [f"Sheet: {sheet_name}", " | ".join(header_cells)]
            for row in group:
                cells = [str(row[i]) if row[i] is not None else "" for i in col_indices]
                lines.append(" | ".join(cells))
            content = "\n".join(lines)
            chunk_id = hashlib.sha256(f"{doc_id}:{content}".encode()).hexdigest()
            chunks.append(Chunk(
                id=chunk_id,
                content=content,
                provenance=ChunkProvenance(
                    document_id=doc_id,
                    source_url=source_url,
                    section_heading=sheet_name,
                    page_number=sheet_idx,
                ),
            ))

    wb.close()
    return source, chunks
